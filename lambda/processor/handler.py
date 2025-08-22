import os
import io
import csv
import json
import boto3
from botocore.client import Config
from decimal import Decimal, InvalidOperation
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime, timezone
import re

s3 = boto3.client("s3", config=Config(signature_version="s3v4"))
dynamodb = boto3.resource("dynamodb")

BUCKET = os.environ["BUCKET_NAME"]
SOURCE_PREFIX = os.environ.get("SOURCE_PREFIX", "source/")
OUTPUT_PREFIX = os.environ.get("OUTPUT_PREFIX", "output/")
DEFAULT_TARGET_TEMPLATE = os.environ.get("TARGET_TEMPLATE_KEY", f"{SOURCE_PREFIX.rstrip('/')}/nasmatch-portfolio.xlsx")
DEFAULT_OUTPUT_KEY = os.environ.get("OUTPUT_KEY_TEMPLATE", f"{OUTPUT_PREFIX.rstrip('/')}/nasmatch-portfolio-updated.xlsx")
MAPPING_TABLE_NAME = os.environ.get("MAPPING_TABLE")  # optional, for account mapping
DEFAULT_DATASET_ID = os.environ.get("DEFAULT_DATASET_ID", "default")
ACCOUNT_NAME_MAP_JSON = os.environ.get("ACCOUNT_NAME_MAP_JSON", "{}")  # fallback JSON mapping
COST_MODE = os.environ.get("COST_MODE", "total_basis").lower()

CSV_COLS = {
    "account_number": "Account Number",
    "account_name":   "Account Name",
    "symbol":         "Symbol",
    "description":    "Description",
    "quantity":       "Quantity",
    "cost_total":     "Cost Basis Total",
    "cost_avg":       "Average Cost Basis",
}

# Consolidate headers
CONSOLIDATE_HEADERS = [
    "Account Number",
    "Account Name",
    "Symbol",
    "Quantity",
    "Cost Basis Total",
    "Average Cost Basis",
    "Type",
]

# CSV required columns
CSV_REQUIRED = {
    "account_number": "Account Number",
    "account_name": "Account Name",
    "symbol": "Symbol",
    "quantity": "Quantity",
    "cost_total": "Cost Basis Total",
    "avg_cost": "Average Cost Basis",
    "type": "Type",
}

# Per-ticker write targets (1-based Excel rows)
ROW_QTY = 24
ROW_COST = 39

def _norm(s):
    return "".join(str(s).strip().lower().split()) if s is not None else ""

def _find_columns(csv_headers):
    norm_to_actual = {_norm(h): h for h in csv_headers}
    resolved, missing = {}, []
    for _, expect in CSV_REQUIRED.items():
        key = _norm(expect)
        hit = norm_to_actual.get(key)
        if not hit:
            missing.append(expect)
        else:
            # back-map with our canonical key
            can_key = [k for k, v in CSV_REQUIRED.items() if v == expect][0]
            resolved[can_key] = hit
    if missing:
        raise ValueError(f"Missing required CSV columns: {missing}. Found headers: {list(csv_headers)}")
    return resolved

def _get_obj_bytes(bucket: str, key: str) -> bytes:
    resp = s3.get_object(Bucket=bucket, Key=key)
    return resp["Body"].read()

def _put_obj_bytes(bucket: str, key: str, body: bytes, content_type: str):
    s3.put_object(Bucket=bucket, Key=key, Body=body, ContentType=content_type)

def _clean_money(val):
    if val is None:
        return None
    if isinstance(val, (int, float, Decimal)):
        return Decimal(str(val))
    s = str(val).strip()
    if not s or s.upper() == "NAN" or s == "-":
        return None
    s = s.replace("$", "").replace(",", "").replace("(", "-").replace(")", "")
    try:
        return Decimal(s)
    except InvalidOperation:
        return None

def _clean_number(val):
    if val is None:
        return None
    if isinstance(val, (int, float, Decimal)):
        return Decimal(str(val))
    s = str(val).strip()
    if not s or s.upper() == "NAN" or s == "-":
        return None
    s = s.replace(",", "")
    try:
        return Decimal(s)
    except InvalidOperation:
        return None

def _get_account_mapping(dataset_id: str):
    # 1) DynamoDB MappingOverrides (PK=dataset_id, SK=source_col -> target_col)
    ddb_map = {}
    if MAPPING_TABLE_NAME:
        tbl = dynamodb.Table(MAPPING_TABLE_NAME)
        try:
            from boto3.dynamodb.conditions import Key as DdbKey
            resp = tbl.query(KeyConditionExpression=DdbKey("dataset_id").eq(dataset_id))
            for it in resp.get("Items", []):
                source_col = it.get("source_col")
                target_col = it.get("target_col") or it.get("target") or source_col
                if source_col and target_col:
                    ddb_map[source_col] = target_col
        except Exception as e:
            print(f"[WARN] DDB mapping query failed: {e}")

    # 2) Fallback env JSON mapping
    env_map = {}
    try:
        env_map = json.loads(ACCOUNT_NAME_MAP_JSON or "{}")
    except Exception as e:
        print(f"[WARN] ACCOUNT_NAME_MAP_JSON parse failed: {e}")

    # Merge (DDB wins)
    merged = {**env_map, **ddb_map}
    print(f"[INFO] Account mapping loaded: {merged}")
    return merged

def _sheet_headers(ws, row=1, start_col=1):
    headers = {}
    c = start_col
    while True:
        val = ws.cell(row=row, column=c).value
        if val is None and c > start_col + 50:  # stop after a reasonable span
            break
        if isinstance(val, str) and val.strip():
            headers[val.strip()] = c
        elif isinstance(val, (int, float)):
            headers[str(val)] = c
        c += 1
    return headers

def _read_positions_csv(csv_bytes):
    """
    Reads the CSV using headers. Expects headers like:
    Account Number,Account Name,Symbol,Description,Quantity,Last Price,Last Price Change,
    Current Value,Today's Gain/Loss Dollar,Today's Gain/Loss Percent,Total Gain/Loss Dollar,
    Total Gain/Loss Percent,Percent Of Account,Cost Basis Total,Average Cost Basis,Type
    """
    import csv, io
    text = csv_bytes.decode('utf-8-sig')
    reader = csv.DictReader(io.StringIO(text))
    rows = []
    missing = [c for c in CSV_COLS.values() if c not in (reader.fieldnames or [])]
    if missing:
        print(f"[WARN] Missing expected columns: {missing}. Proceeding with what exists.")

    for r in reader:
        acct_name = (r.get(CSV_COLS["account_name"]) or "").strip()
        sym = (r.get(CSV_COLS["symbol"]) or "").strip().upper()
        if not acct_name or not sym:
            continue

        qty_raw = r.get(CSV_COLS["quantity"])
        qty = _to_float(qty_raw)

        if COST_MODE == "avg_per_share":
            cost_raw = r.get(CSV_COLS["cost_avg"])
        else:
            cost_raw = r.get(CSV_COLS["cost_total"])
        cost = _to_money(cost_raw)

        rows.append({
            "account_name": acct_name,
            "symbol": sym,
            "qty": qty,
            "cost": cost,
        })

    print(f"[PARSE] Parsed {len(rows)} position rows (COST_MODE={COST_MODE})")
    return rows

def _ensure_consolidate_sheet(wb):
    if "consolidate" in wb.sheetnames:
        ws = wb["consolidate"]
    else:
        ws = wb.create_sheet("consolidate")
    for idx, h in enumerate(CONSOLIDATE_HEADERS, start=1):
        ws.cell(row=1, column=idx, value=h)
    return ws

def _clear_data_rows(ws, start_row=2):
    max_row = ws.max_row
    if max_row >= start_row:
        ws.delete_rows(idx=start_row, amount=max_row - start_row + 1)

def _write_consolidate(ws, rows):
    _clear_data_rows(ws, start_row=2)
    for r_idx, item in enumerate(rows, start=2):
        for c_idx, h in enumerate(CONSOLIDATE_HEADERS, start=1):
            ws.cell(row=r_idx, column=c_idx, value=item.get(h))
    # number formats
    hdr_index = {h: i+1 for i, h in enumerate(CONSOLIDATE_HEADERS)}
    if "Quantity" in hdr_index:
        col = hdr_index["Quantity"]
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=col)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.######"
    for money_col in ("Cost Basis Total", "Average Cost Basis"):
        if money_col in hdr_index:
            col = hdr_index[money_col]
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(row=r, column=col)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '"$"#,##0.00;[Red]-"$"#,##0.00'

def _load_workbook_from_s3(key: str):
    try:
        xbytes = _get_obj_bytes(BUCKET, key)
        return load_workbook(io.BytesIO(xbytes))
    except Exception as e:
        print(f"[INFO] Could not load workbook at {key} ({e}); creating a new workbook.")
        return Workbook()

def _symbol_from_sheetname(name: str):
    """Allow sheet names like 'AAPL', 'AAPL (summary)', etc."""
    base = name.strip().upper()
    if " " in base:
        base = base.split(" ")[0]
    if "(" in base:
        base = base.split("(")[0]
    return base.strip()

def _aggregate_by_symbol_account(rows, acct_map):
    """
    Build: agg[symbol][account_column] = { qty: sum, cost: sum }
    'account_column' is the mapped target column header from sheet row 1 (e.g., 'BrokerageLink').
    """
    agg = {}
    for r in rows:
        symbol = (r.get("Symbol") or "").strip().upper()
        if not symbol:
            continue
        src_acct = (r.get("Account Name") or "").strip()
        if not src_acct:
            continue
        tgt_acct = acct_map.get(src_acct)  # map to sheet header
        if not tgt_acct:
            # unmapped account gets skipped for per-ticker write
            continue
        qty = r.get("Quantity") or 0.0
        cost = r.get("Cost Basis Total") or 0.0
        agg.setdefault(symbol, {}).setdefault(tgt_acct, {"qty": 0.0, "cost": 0.0})
        agg[symbol][tgt_acct]["qty"] += float(qty or 0.0)
        agg[symbol][tgt_acct]["cost"] += float(cost or 0.0)
    return agg

def _write_per_ticker(wb, agg):
    """
    Returns:
      updated_count, headers_by_sheet, missing_accounts, writes_detail
    """
    updated = 0
    headers_by_sheet = {}
    missing_accounts = {}
    writes_detail = {}  # {symbol: {account: {qty, cost, qty_written, cost_written}}}

    for sheetname in wb.sheetnames:
        sym = _symbol_from_sheetname(sheetname)
        if sym not in agg:
            continue
        ws = wb[sheetname]
        headers = _sheet_headers(ws, row=1, start_col=2)
        headers_by_sheet[sheetname] = list(headers.keys())
        sym_writes = writes_detail.setdefault(sym, {})

        for acct_name, sums in agg[sym].items():
            col = None
            # case/space-insensitive match
            normalized = { _norm(h): c for h, c in headers.items() }
            col = normalized.get(_norm(acct_name))
            if not col:
                missing_accounts.setdefault(sheetname, []).append(acct_name)
                print(f"[INFO] Sheet '{sheetname}': account '{acct_name}' not found among headers {list(headers.keys())}")
                continue

            # write qty (row 24) and cost (row 39)
            ws.cell(row=ROW_QTY, column=col, value=sums["qty"])
            ws.cell(row=ROW_QTY, column=col).number_format = "0.######"
            ws.cell(row=ROW_COST, column=col, value=sums["cost"])
            ws.cell(row=ROW_COST, column=col).number_format = '"$"#,##0.00;[Red]-"$"#,##0.00'
            updated += 1

            sym_writes[acct_name] = {
                "qty": sums["qty"],
                "cost": sums["cost"],
                "qty_written": sums["qty"],
                "cost_written": sums["cost"]
            }
    return updated, headers_by_sheet, missing_accounts, writes_detail

def _process_all(source_key: str, target_key: str, output_key: str):
    csv_bytes = _get_obj_bytes(BUCKET, source_key)
    rows = _read_positions_csv(csv_bytes)

    wb = _load_workbook_from_s3(target_key)

    # 1) consolidate
    ws = _ensure_consolidate_sheet(wb)
    _write_consolidate(ws, rows)

    # 2) per-ticker
    acct_map = _get_account_mapping(DEFAULT_DATASET_ID)
    agg = _aggregate_by_symbol_account(rows, acct_map)
    updated, headers_by_sheet, missing_accounts, writes_detail = _write_per_ticker(wb, agg)
    print(f"[INFO] Per-ticker writes: {updated}")

    # 3) report objects
    ts = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    json_key = f"{OUTPUT_PREFIX.rstrip('/')}/reports/{ts}-report.json"
    report = _build_report(rows, agg, headers_by_sheet, writes_detail, missing_accounts)
    _write_runreport_sheet(wb, report, source_key, target_key, output_key)
    _save_json_report(BUCKET, json_key, report)

    # 4) save workbook
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    _put_obj_bytes(
        BUCKET, output_key, buf.read(),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    return {
        "status": "ok",
        "bucket": BUCKET,
        "output_key": output_key,
        "report_key": json_key,
        "wrote_rows": len(rows),
        "per_ticker_writes": updated,
        "symbols_updated": report["symbols_updated"],
        "missing_accounts": missing_accounts
    }

def _ts():
    return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S %Z")

def _save_json_report(bucket, key, payload: dict):
    s3.put_object(
        Bucket=bucket,
        Key=key,
        Body=json.dumps(payload, indent=2).encode("utf-8"),
        ContentType="application/json"
    )
    print(f"[REPORT] Wrote JSON report to s3://{bucket}/{key}")

def _build_report(rows, agg, headers_by_sheet, writes_detail, missing_accounts):
    symbols_updated = sorted(list(writes_detail.keys()))
    total_writes = sum(
        sum(1 for _acct, vals in acct_map.items() if (vals.get("qty_written") or vals.get("cost_written")))
        for acct_map in writes_detail.values()
    )
    return {
        "timestamp": _ts(),
        "summary": {
            "csv_rows_parsed": len(rows),
            "symbols_updated_count": len(symbols_updated),
            "per_ticker_writes": total_writes
        },
        "symbols_updated": symbols_updated,
        "headers_by_sheet": headers_by_sheet,        # {sheet: ["BrokerageLink", ...]}
        "missing_accounts": missing_accounts,        # {sheet: ["Acct not found", ...]}
        "writes_detail": writes_detail               # {symbol: {account: {qty, cost, qty_written, cost_written}}}
    }

def _write_runreport_sheet(wb, report, src_key, tgt_key, out_key):
    wsname = "RunReport"
    if wsname in wb.sheetnames:
        ws = wb[wsname]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(wsname, 0)

    rows = [
        ["RunReport generated", report["timestamp"]],
        ["Source CSV", src_key],
        ["Target Template", tgt_key],
        ["Output Excel", out_key],
        ["CSV Rows Parsed", report["summary"]["csv_rows_parsed"]],
        ["Symbols Updated", report["summary"]["symbols_updated_count"]],
        ["Per-ticker Writes", report["summary"]["per_ticker_writes"]],
        [],
        ["Symbol", "Account Header", "Qty Written", "Cost Written"]
    ]
    for sym in report["symbols_updated"]:
        acct_map = report["writes_detail"].get(sym, {})
        for acct, vals in acct_map.items():
            rows.append([sym, acct, vals.get("qty_written") or 0, vals.get("cost_written") or 0])

    # Missing accounts block
    rows.append([])
    rows.append(["Sheet", "Missing account headers (not found in row 1)"])
    for sheet, miss in (report.get("missing_accounts") or {}).items():
        rows.append([sheet, ", ".join(miss)])

    for r in rows:
        ws.append(r)
    ws.freeze_panes = "A9"

def _to_money(s):
    if s is None: return 0.0
    if isinstance(s, (int, float)): return float(s)
    s = str(s).strip()
    s = s.replace(',', '')
    m = re.sub(r'[^\d\.\-\+]', '', s)  # keep digits, dot, sign
    try: return float(m) if m else 0.0
    except: return 0.0

def _to_float(s):
    if s is None: return 0.0
    try: return float(str(s).replace(',', '').strip())
    except: return 0.0



def main(event, context):
    global BUCKET  # <-- declare first, before BUCKET is referenced anywhere
    print("Event:", json.dumps(event))

    # S3 event path
    if isinstance(event, dict) and event.get("Records"):
        rec = event["Records"][0]
        s3info = rec.get("s3", {})
        bucket = s3info.get("bucket", {}).get("name", BUCKET)
        key = s3info.get("object", {}).get("key")
        if not key:
            raise ValueError("S3 event missing object key")
        if not key.lower().endswith(".csv"):
            return {"status": "ignored", "reason": "not a .csv", "key": key}

        # respect event bucket (multi-bucket scenarios)
        BUCKET = bucket

        target_key = DEFAULT_TARGET_TEMPLATE
        output_key = DEFAULT_OUTPUT_KEY
        if not target_key.startswith(SOURCE_PREFIX) and not target_key.startswith(OUTPUT_PREFIX):
            target_key = f"{SOURCE_PREFIX}{target_key}"
        if not output_key.startswith(OUTPUT_PREFIX):
            output_key = f"{OUTPUT_PREFIX}{output_key}"
        return _process_all(key, target_key, output_key)

    # Direct JSON path
    source_key = event.get("source_key")
    target_key = event.get("target_key") or DEFAULT_TARGET_TEMPLATE
    output_key = event.get("output_key") or DEFAULT_OUTPUT_KEY
    if not source_key or not source_key.lower().endswith(".csv"):
        raise ValueError("source_key is required and must be a .csv")
    if not source_key.startswith(SOURCE_PREFIX):
        source_key = f"{SOURCE_PREFIX}{source_key}"
    if not target_key.startswith(SOURCE_PREFIX) and not target_key.startswith(OUTPUT_PREFIX):
        target_key = f"{SOURCE_PREFIX}{target_key}"
    if not output_key.startswith(OUTPUT_PREFIX):
        output_key = f"{OUTPUT_PREFIX}{output_key}"
    return _process_all(source_key, target_key, output_key)

