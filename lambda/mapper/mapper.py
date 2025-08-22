import os
import io
import json
import csv
import math
import boto3
from botocore.client import Config
from openpyxl import load_workbook

s3 = boto3.client("s3", config=Config(signature_version="s3v4"))
dynamodb = boto3.resource("dynamodb")
bedrock = boto3.client("bedrock-runtime")

BUCKET = os.environ["BUCKET_NAME"]
MAPPING_TABLE = os.environ["MAPPING_TABLE"]
DATASET_ID = os.environ.get("DEFAULT_DATASET_ID", "default")
TEXT_MODEL = os.environ["BEDROCK_TEXT_MODEL_ID"]
EMBED_MODEL = os.environ["BEDROCK_EMBED_MODEL_ID"]

def _get_obj_bytes(key):
    return s3.get_object(Bucket=BUCKET, Key=key)["Body"].read()

def _read_csv_headers_and_accounts(csv_key):
    b = _get_obj_bytes(csv_key)
    text = b.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    headers = reader.fieldnames or []
    # grab distinct account names (some rows may be blank)
    accounts = []
    for r in reader:
        v = (r.get("Account Name") or "").strip()
        if v and v not in accounts:
            accounts.append(v)
            if len(accounts) > 2000:  # guard
                break
    return headers, accounts

def _read_excel_account_headers(template_key):
    b = _get_obj_bytes(template_key)
    wb = load_workbook(io.BytesIO(b))
    # read row1 from the first ticker-like sheet that has account columns in B.. (fallback to 'consolidate')
    headers = set()
    for name in wb.sheetnames:
        ws = wb[name]
        # read row1 B..Z for candidate headers
        for col in range(2, 60):
            val = ws.cell(row=1, column=col).value
            if isinstance(val, str) and val.strip():
                headers.add(val.strip())
        # heuristic: if we found candidates on this sheet, good enough
        if headers:
            break
    if not headers and "consolidate" in wb.sheetnames:
        ws = wb["consolidate"]
        for col in range(2, 60):
            val = ws.cell(row=1, column=col).value
            if isinstance(val, str) and val.strip():
                headers.add(val.strip())
    return sorted(headers)

def _embed(texts):
    body = json.dumps({
        "inputText": texts if isinstance(texts, str) else " ".join(texts)
    })
    # Titan V2 returns embedding under "embedding"
    resp = bedrock.invoke_model(
        modelId=EMBED_MODEL,
        body=body,
        contentType="application/json",
        accept="application/json",
    )
    out = json.loads(resp["body"].read())
    if "embedding" in out:
        return out["embedding"]
    # Some variants return list of embeddings
    return out.get("embeddings", [])[0]

def _cosine(a, b):
    num = sum(x*y for x, y in zip(a, b))
    da = math.sqrt(sum(x*x for x in a))
    db = math.sqrt(sum(y*y for y in b))
    return num / (da*db + 1e-9)

def _rank_by_similarity(src, candidates):
    e_src = _embed(src)
    ranks = []
    for c in candidates:
        e_c = _embed(c)
        ranks.append((c, _cosine(e_src, e_c)))
    ranks.sort(key=lambda t: t[1], reverse=True)
    return ranks

def _llm_decide_mapping(account_src, ranked_candidates):
    """
    Few-shot prompt to choose best candidate.
    """
    top = [f"{name} (score={score:.3f})" for name, score in ranked_candidates[:5]]
    prompt = f"""
You are normalizing account names across brokerage data.

Given an incoming Account Name from a CSV: "{account_src}"
and a set of candidate Excel account headers:

{top}

Pick the single best target header. If none fit semantically, reply with EXACTLY: "UNMAPPED".

Rules:
- Prefer semantic match over string similarity.
- "401K" should map to "BrokerageLink" (Fidelity 401k brokerage window).
- "401K Roth" should map to "BrokerageLink Roth".
- Return ONLY the header text, no extra words, unless UNMAPPED.
"""
    body = json.dumps({
        "anthropic_version": "bedrock-2023-05-31",
        "max_tokens": 128,
        "messages": [{"role": "user", "content": prompt}]
    })
    resp = bedrock.invoke_model(
        modelId=TEXT_MODEL,
        body=body,
        contentType="application/json",
        accept="application/json",
    )
    result = json.loads(resp["body"].read())
    # Claude on Bedrock returns choices under content[0].text
    txt = ""
    try:
        txt = result["content"][0]["text"].strip()
    except Exception:
        txt = ""
    if not txt:
        return "UNMAPPED"
    # sanitize common quotes
    return txt.strip().strip('"').strip("'")

def _save_mapping(dataset_id, source_col, target_col):
    tbl = dynamodb.Table(MAPPING_TABLE)
    tbl.put_item(Item={
        "dataset_id": dataset_id,
        "source_col": source_col,
        "target_col": target_col
    })

def _load_account_map():
    env_map = json.loads(os.environ.get("ACCOUNT_NAME_MAP_JSON", "{}") or "{}")
    # normalize keys and values
    return { _norm(k): v for k, v in env_map.items() }

def _apply_mapping(account_name: str, ddb_map: dict, env_map: dict) -> str | None:
    # 1) DDB override
    v = ddb_map.get(_norm(account_name))
    if v: return v
    # 2) Env fallback
    return env_map.get(_norm(account_name))
    
def main(event, context):
    print("Event:", json.dumps(event))
    csv_key = event.get("csv_key")
    template_key = event.get("template_key")
    dataset_id = event.get("dataset_id", DATASET_ID)
    if not csv_key or not template_key:
        return {"ok": False, "error": "csv_key and template_key are required"}

    csv_headers, csv_accounts = _read_csv_headers_and_accounts(csv_key)
    excel_headers = _read_excel_account_headers(template_key)
    if not excel_headers:
        return {"ok": False, "error": "No account headers found in Excel row1 (B..)"}

    created = 0
    decisions = []
    for acct in csv_accounts:
        # rank by embeddings
        ranks = _rank_by_similarity(acct, excel_headers)
        # ask LLM to choose (with domain rules)
        choice = _llm_decide_mapping(acct, ranks)
        if choice != "UNMAPPED":
            _save_mapping(dataset_id, acct, choice)
            created += 1
        decisions.append({"source": acct, "choice": choice, "candidates": ranks[:3]})
    return {
        "ok": True,
        "dataset_id": dataset_id,
        "created_mappings": created,
        "excel_headers_considered": excel_headers,
        "decisions": decisions[:20]  # truncate for payload size
    }
